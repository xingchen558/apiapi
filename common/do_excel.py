# -*- coding: utf-8 -*-
import openpyxl
import os
from common import contants
from common.request import Request
import json
class Case:
    def __init__(self):
        self.case_id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
class DoExcel:
    def __init__(self, file_name):
        try:
            self.file_name = file_name  # 保存的时候，会调用excel名称
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except Exception as e:
            print("{} is not found!".format(file_name))
            raise e
    def get_cases(self, sheet_name):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        cases = []
        for i in range(2, max_row+1):
            case = Case()
            case.case_id = sheet.cell(row=i, column=1).value
            case.url = sheet.cell(row=i, column=2).value
            case.data = sheet.cell(row=i, column=3).value
            case.title = sheet.cell(row=i, column=4).value
            case.method = sheet.cell(row=i, column=5).value
            case.expected = sheet.cell(row=i, column=6).value
            cases.append(case)
        return cases
    def get_sheet_names(self):  # 获取到所有的sheet名称的列表
        return self.workbook.sheetnames
    # 根据sheet_name定位到sheet， 然后根据cased_id定位到行，渠道当前行里面actual这个单元格
    def write_actual_by_case_id(self, sheet_name, case_id, actual):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for i in range(2, max_row+1):
            case_id_i = sheet.cell(i, 1).value
            if case_id_i == case_id:  # 判断excel里面的当前行的case_id是否等于传入的case_id
                sheet.cell(i, 7).value = actual  # 写入传进来的actual到excel中actual单元格内
                self.workbook.save(filename=self.file_name)  # 保存文档
                break  # 一直遍历，必须中断
    def write_result_by_case_id(self, sheet_name, case_id, result):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for i in range(2, max_row+1):
            case_id_i = sheet.cell(i, 1).value
            if case_id_i == case_id:  # 判断excel里面的当前行的case_id是否等于传入的case_id
                sheet.cell(i, 8).value = result  # 写入传进来的actual到excel中actual单元格内
                self.workbook.save(filename=self.file_name)  # 保存文档
                print(result)
                break

if __name__ == '__main__':
    file_name = os.path.join(contants.datas_dir, 'test_cases.xlsx')
    do_excel = DoExcel(file_name)
    sheet_names = do_excel.get_sheet_names()  # 获取到所有的sheet名称的列表

    print("sheet 名称列表：", sheet_names)
    case_list = ["register", "login"]  # 定义执行的sheet列表
    for sheet_name in sheet_names:
        if sheet_name in case_list:  # 如果sheet_name在list中，就执行
            cases = do_excel.get_cases(sheet_name)
            print(sheet_name + "测试用例个数：", len(cases))
            for case in cases:  # 遍历测试用例列表，for一次，则去一个case实例
                print('cases:', case.__dict__)  # 打印case信息
                data = eval(case.data)  # 渠道data字符串，然后转化成字典
                res = Request(method=case.method, url=case.url, data=data)  # 接口调用
                print("status_code:", res.get_status_code())
                res_dict = res.get_json()  # 获取请求响应：字典
                # 把字段转化成格式化的字符串
                res_text = json.dumps(res_dict, ensure_ascii=False, indent=4)
                print("response:", res_text)  # 打印响应
                do_excel.write_actual_by_case_id(sheet_name=sheet_name, case_id=case.case_id, actual=res.get_text())

                if case.expected == res.get_text():
                    print("result : PASS")
                    do_excel.write_result_by_case_id(sheet_name=sheet_name, case_id=case.case_id, result='PASS')
                else:
                    print("result : FAIL")
                    do_excel.write_result_by_case_id(sheet_name=sheet_name, case_id=case.case_id, result='FAIL')











